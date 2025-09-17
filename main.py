import time
import sys
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

SUBJECT_ROW: int = 2
SCORE_ROW: int = 4
PARALLELL_COLUMN: int = 4
SHEET_DEFAULT_NAME: str = "school_marks_count"


def save_as_xlsx(data: list) -> None:
    """Save as .xlsx file with headers."""

    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "district",
            "organisation",
            "parallel",
            "subject",
            "score",
            "count",
        ]
    )

    [ws.append(row) for row in data]
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    filename = int(time.time())
    path = Path("output")
    path.mkdir(parents=True, exist_ok=True)

    wb.save(path / f"{filename}.xlsx")


def map_values(
    ws: Worksheet,
    subjects: list[str],
    scores: list[int],
) -> list[list[str, str, int, str, int]]:
    """
    Get complete list of pairs district -> organisation
    of every row after the header.
    """

    result = []
    bad_scores_row = list(ws.rows)[SCORE_ROW - 1]

    # since we start right after header
    for row in ws.iter_rows(
        values_only=True,
        min_row=SCORE_ROW + 1,
    ):
        district = row[0]
        organisation = row[1]
        for score, subject in zip(scores, subjects):
            row_count = row[score]
            row_subject = subject
            row_score = bad_scores_row[score - 1].value
            # zero based since not a row object
            row_parallell = row[PARALLELL_COLUMN - 1]
            if row_count > 0:
                result.append(
                    [
                        district,
                        organisation,
                        row_parallell,
                        row_subject,
                        row_score,
                        row_count,
                    ]
                )

    return result


def bad_subjects(ws: Worksheet, columns: list[int]) -> list[str]:
    """
    Iterate over subject row and get subjects
    that are direct parents of invalid scores.
    """
    subjects = []

    for column in columns:
        for row in ws.iter_rows(
            values_only=True,
            min_row=SUBJECT_ROW,
            max_row=SUBJECT_ROW,
            min_col=column,
            max_col=column,
        ):
            for i, cell in enumerate(row, 1):
                subjects.append(cell)

    return subjects


def bad_score_columns(ws: Worksheet) -> list[int]:
    """
    Iterate over scores row and get scores columns
    that are invalid (has math operators).
    """
    columns = []

    for row in ws.iter_rows(
        values_only=True,
        min_row=SCORE_ROW,
        max_row=SCORE_ROW,
    ):
        for column, cell in enumerate(row, 1):
            if cell is not None:
                if "+" in cell or "-" in cell:
                    columns.append(column)

    return columns


def main():
    if len(sys.argv) > 1:
        filename = sys.argv[1]
    else:
        filename = "input.xlsx"

    wb = load_workbook(filename, read_only=True)
    ws = wb[SHEET_DEFAULT_NAME]

    columns = bad_score_columns(ws)
    subjects = bad_subjects(ws, columns)
    data = map_values(ws, subjects, columns)
    save_as_xlsx(data)


if __name__ == "__main__":
    main()
